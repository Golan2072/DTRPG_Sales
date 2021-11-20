# Stellagama Publishing Account Analysis Software

import csv
import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import Font


class Book_product:
    def __init__(self, product_name):
        self.name = product_name
        self.total_count = 0
        self.monthly_count = 0
        self.total_revenue = 0
        self.monthly_revenue = 0
        self.author = ""
        self.cost = 0
        self.release_month = 0
        self.release_year = 0
        # royalties
        self.richard = 0
        self.josh = 0
        self.bob = 0
        self.hannah = 0
        self.ivan = 0


class Data_item:
    def __init__(self, data_name):
        self.name = data_name
        self.author = ""
        self.cost = 0
        self.release_month = 0
        self.release_year = 0


def month_calculation(start_month, start_year, end_month, end_year):
    year_difference = int(end_year) - int(start_year)
    if year_difference < 0:
        year_difference = 0
    return year_difference * 12 + int(end_month) - int (start_month)


def product_lister(dict_list):
    products = []
    for dict in dict_list:
        if dict["Name"] not in products:
            products.append(dict["Name"])
    products.sort()
    products.remove("")
    return products


def merge_products(main_product, secondary_product):
    new_product = Book_product(main_product)
    new_product.name = main_product.name
    new_product.total_count = main_product.total_count + secondary_product.total_count
    new_product.monthly_count = main_product.monthly_count + secondary_product.monthly_count
    new_product.total_revenue = main_product.total_revenue + secondary_product.total_revenue
    new_product.monthly_revenue = main_product.monthly_revenue + secondary_product.monthly_revenue
    return new_product


def product_cleaner(product_dictionary):
    output_product_dictionary = {}
    for product in product_dictionary:
        if product_dictionary[product].total_revenue != 0:
            output_product_dictionary[product] = product_dictionary[product]
        else:
            pass
    output_product_dictionary["TSAO: Wreck in the Ring"] = merge_products(product_dictionary["TSAO: Wreck in the Ring"],
                                                                          product_dictionary[
                                                                              "Borderlands Adventure 1: Wreck in the Ring"])
    output_product_dictionary.pop("Borderlands Adventure 1: Wreck in the Ring")
    output_product_dictionary["Character Options for Stars Without Number"] = merge_products(
        product_dictionary["Character Options for Stars Without Number"],
        product_dictionary["Character Options - compatible with Stars Without Number"])
    output_product_dictionary.pop("Character Options - compatible with Stars Without Number")
    output_product_dictionary["Cheating Death 2nd Edition"] = merge_products(
        product_dictionary["Cheating Death 2nd Edition"], product_dictionary["Cheating Death"])
    output_product_dictionary.pop("Cheating Death")
    output_product_dictionary["From the Ashes 2nd Edition"] = merge_products(
        product_dictionary["From the Ashes 2nd Edition"], product_dictionary["From the Ashes"])
    output_product_dictionary.pop("From the Ashes")
    output_product_dictionary["TSAO: Stationery and Heraldry Pack"] = merge_products(
        product_dictionary["TSAO: Stationery and Heraldry Pack"],
        product_dictionary["TSAO: Stationary and Heraldry Pack"])
    output_product_dictionary.pop("TSAO: Stationary and Heraldry Pack")
    output_product_dictionary["TSAO: Liberty Ship"] = merge_products(product_dictionary["TSAO: Liberty Ship"],
                                                                     product_dictionary["Liberty Ship"])
    output_product_dictionary.pop("Liberty Ship")
    output_product_dictionary["TSAO: 50 Wonders of the Reticulan Empire"] = merge_products(
        product_dictionary["TSAO: 50 Wonders of the Reticulan Empire"],
        product_dictionary["50 Wonders of the Reticulan Empire"])
    output_product_dictionary.pop("50 Wonders of the Reticulan Empire")
    output_product_dictionary["The Sword of Cepheus"] = merge_products(product_dictionary["The Sword of Cepheus"],
                                                                       product_dictionary["Sword of Cepheus"])
    output_product_dictionary.pop("Sword of Cepheus")
    output_product_dictionary["TSAO: Wreck in the Ring"] = merge_products(product_dictionary["TSAO: Wreck in the Ring"],
                                                                          product_dictionary[
                                                                              "TSAO: Borderlands Adventure 1: Wreck in the Ring"])
    output_product_dictionary.pop("TSAO: Borderlands Adventure 1: Wreck in the Ring")
    output_product_dictionary["TSAO: These Stars Are Ours!"] = merge_products(
        product_dictionary["TSAO: These Stars Are Ours!"], product_dictionary["These Stars Are Ours!"])
    output_product_dictionary.pop("These Stars Are Ours!")
    output_product_dictionary["Barbaro!"] = output_product_dictionary["Â¡BÃ¡rbaro!"]
    output_product_dictionary.pop("Â¡BÃ¡rbaro!")
    return output_product_dictionary


def excel_output(product_dictionary, month, year):
    filename = "sales-" + month + year + ".xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Monthly Sales Report {month}{year}"
    sheet['A1'] = f"Monthly Sales Report {month}/{year}"
    sheet['A1'].font = Font(size=16, bold=True)
    for cell in ["A3", "B3", "C3", "D3", "E3", "F3", "G3", "H3", "I3", "J3", "K3", "L3", "M3", "N3", "A60", "F60",
                 "G60", "N60", "O3", "P3", "Q3", "R3", "S3", "O60", "P60", "Q60", "R60", "S60"]:
        sheet[cell].font = Font(bold=True)
    sheet['A3'] = "Product"
    sheet['B3'] = "Author"
    sheet['C3'] = "Production Cost"
    sheet['D3'] = "Total Paid Sales"
    sheet['E3'] = "Monthly Paid Sales"
    sheet['F3'] = "Total Revenue"
    sheet['G3'] = "Monthly Revenue"
    sheet['H3'] = "Release Year"
    sheet['I3'] = "Release Month"
    sheet['J3'] = "Months Since Release"
    sheet['K3'] = "Net Profit"
    sheet['L3'] = "Avg. Monthly Sales"
    sheet['M3'] = "Avg. Monthly Profit"
    sheet['N3'] = "Owner's Share"
    sheet['O3'] = "Richard's Royalties"
    sheet['P3'] = "Josh's Royalties"
    sheet['Q3'] = "Bob's Royalties"
    sheet['R3'] = "Hannah's Royalties"
    sheet['S3'] = "Ivan's Royalties"
    sheet['A60'] = "Total"
    sheet["F60"] = "=SUM(F4:F59)"
    sheet["G60"] = "=SUM(G4:G59)"
    sheet["N60"] = "=SUM(N4:N59)"
    sheet["O60"] = "=SUM(O4:O59)"
    sheet["P60"] = "=SUM(P4:P59)"
    sheet["Q60"] = "=SUM(Q4:Q59)"
    sheet["R60"] = "=SUM(R4:R59)"
    sheet["S60"] = "=SUM(S4:S59)"
    for row in sheet.iter_rows(min_row=2, max_row=1000):
        for column in sheet.iter_cols(min_row=2, max_row=150, min_col=6, max_col=7):
            for cell in column:
                cell.number_format = '$#,##0.00'
        for column in sheet.iter_cols(min_row=2, max_row=150, min_col=3, max_col=3):
            for cell in column:
                cell.number_format = '$#,##0.00'
    for column in sheet.iter_cols(min_row=2, max_row=150, min_col=11, max_col=11):
        for cell in column:
            cell.number_format = '$#,##0.00'
    for column in sheet.iter_cols(min_row=2, max_row=150, min_col=12, max_col=12):
        for cell in column:
            cell.number_format = '0.00'
    for column in sheet.iter_cols(min_row=2, max_row=150, min_col=13, max_col=19):
        for cell in column:
            cell.number_format = '$#,##0.00'
    sheet.column_dimensions['A'].width = 52
    sheet.column_dimensions['B'].width = 32
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20
    sheet.column_dimensions['M'].width = 20
    sheet.column_dimensions['N'].width = 20
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['P'].width = 20
    sheet.column_dimensions['Q'].width = 20
    sheet.column_dimensions['R'].width = 20
    sheet.column_dimensions['S'].width = 20
    excel_row = 3
    for product in product_dictionary:
        excel_row += 1
        profit = float(float(product_dictionary[product].total_revenue) - float(product_dictionary[product].cost))
        month_differential = month_calculation(int(product_dictionary[product].release_month),
                                               int(product_dictionary[product].release_year), int(month), int(year))
        if month_differential < 0:
            month_differential = 1
        if month_differential > 0:
            monthly_sales = product_dictionary[product].total_count / month_differential
            if profit < 0:
                monthly_profit = 0
            else:
                monthly_profit = profit / month_differential
        else:
            monthly_sales = product_dictionary[product].total_count
            monthly_profit = profit
            if profit < 0:
                monthly_profit = 0
            else:
                monthly_profit = profit
        if profit > 0:
            owners_share = 0.25 * product_dictionary[product].monthly_revenue
            richard_royalties = float(product_dictionary[product].richard) * float(product_dictionary[product].monthly_revenue)
            josh_royalties = product_dictionary[product].josh * product_dictionary[product].monthly_revenue
            bob_royalties = product_dictionary[product].bob * product_dictionary[product].monthly_revenue
            hannah_royalties = product_dictionary[product].hannah * product_dictionary[product].monthly_revenue
            ivan_royalties = product_dictionary[product].ivan * product_dictionary[product].monthly_revenue
        else:
            owners_share = 0
            richard_royalties = 0
            josh_royalties = 0
            bob_royalties = 0
            hannah_royalties = 0
            ivan_royalties = 0
        sheet.cell(row=excel_row, column=1).value = str(product_dictionary[product].name)
        sheet.cell(row=excel_row, column=2).value = str(product_dictionary[product].author)
        sheet.cell(row=excel_row, column=3).value = float(product_dictionary[product].cost)
        sheet.cell(row=excel_row, column=4).value = float(product_dictionary[product].total_count)
        sheet.cell(row=excel_row, column=5).value = float(product_dictionary[product].monthly_count)
        sheet.cell(row=excel_row, column=6).value = float(product_dictionary[product].total_revenue)
        sheet.cell(row=excel_row, column=7).value = float(product_dictionary[product].monthly_revenue)
        sheet.cell(row=excel_row, column=8).value = int(product_dictionary[product].release_year)
        sheet.cell(row=excel_row, column=9).value = int(product_dictionary[product].release_month)
        if int(year) < int(product_dictionary[product].release_year):
            month_differential = "-"
            sheet.cell(row=excel_row, column=10).value = "-"
            sheet.cell(row=excel_row, column=13).value = "-"
            sheet.cell(row=excel_row, column=11).value = "-"
        else:
            sheet.cell(row=excel_row, column=10).value = month_differential
            sheet.cell(row=excel_row, column=12).value = monthly_sales
            sheet.cell(row=excel_row, column=13).value = monthly_profit
            sheet.cell(row=excel_row, column=11).value = profit
        sheet.cell(row=excel_row, column=14).value = owners_share
        sheet.cell(row=excel_row, column=15).value = richard_royalties
        sheet.cell(row=excel_row, column=16).value = josh_royalties
        sheet.cell(row=excel_row, column=17).value = bob_royalties
        sheet.cell(row=excel_row, column=18).value = hannah_royalties
        sheet.cell(row=excel_row, column=19).value = ivan_royalties
    workbook.save(filename)


if __name__ == '__main__':
    date_month = input("Please input the current month (MM)")
    date_year = input("Please input the current year (YYYY)")
    current_date_1 = date_year + "-" + date_month
    current_date_2 = date_month + "/" + date_year
    with open("dtrpg-report.csv", "r", errors="ignore") as raw_data:
        reader = csv.DictReader(raw_data)
        dict_list = []
        for line in reader:
            dict_list.append(line)
    products = product_lister(dict_list)
    product_dictionary = {}
    for product in products:
        product_dictionary[product] = Book_product(product)
        for product_dict in dict_list:
            if product_dict["Name"] == product:
                product_dictionary[product].total_revenue += float(product_dict["Earnings"])
                product_dictionary[product].total_count += int(product_dict["Quantity"])
                if product_dict["Date"][0:7] == current_date_1 or product_dict["Date"][3:10] == current_date_2:
                    product_dictionary[product].monthly_revenue += float(product_dict["Earnings"])
                    product_dictionary[product].monthly_count += int(product_dict["Quantity"])
                if product_dict["Earnings"] == "0":
                    pass
                else:
                    pass
            else:
                pass
    product_dictionary = product_cleaner(product_dictionary)
    data_list = []
    with open("database.csv", "r", errors="ignore") as base_data:
        reader = csv.DictReader(base_data)
        item_dictionary = {}
        for line in reader:
            if line["ï»¿Product"] != "Ã‚Â¡BÃƒÂ¡rbaro!":
                product_dictionary[line["ï»¿Product"]].cost = line["Cost"]
                product_dictionary[line["ï»¿Product"]].author = line["Author"]
                product_dictionary[line["ï»¿Product"]].release_year = line["Release_Year"]
                product_dictionary[line["ï»¿Product"]].release_month = line["Release_Month"]
                product_dictionary[line["ï»¿Product"]].richard = float(line["Richard"])
                product_dictionary[line["ï»¿Product"]].josh = float(line["Josh"])
                product_dictionary[line["ï»¿Product"]].bob = float(line["Bob"])
                product_dictionary[line["ï»¿Product"]].hannah = float(line["Hannah"])
                product_dictionary[line["ï»¿Product"]].ivan = float(line["Ivan"])
            else:
                pass
    print("Processing...")
    excel_output(product_dictionary, date_month, date_year)
